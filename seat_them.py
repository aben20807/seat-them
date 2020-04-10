from openpyxl import load_workbook
from collections import namedtuple
import argparse

def get_args():
    """ Init argparser and return the args from cli.
    """
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "-s", "--student",
        help="student list",
        type=str,
        default='student.xlsx')

    parser.add_argument(
        "-n", "--number",
        help="the number of the students",
        type=int,
        default=500)

    parser.add_argument(
        "-e", "--seat",
        help="seat with marks",
        type=str,
        default='seat.xlsx')

    parser.add_argument(
        "-o", "--output",
        help="output result",
        type=str,
        default='new_seat.xlsx')
    return parser.parse_args()

def main():
    # Get args
    args = get_args()
    student_filename = args.student
    seat_filename = args.seat
    output_filename = args.output

    # Get and store student
    Student = namedtuple('Student', ('name', 'idnum'))
    stu_wb = load_workbook(student_filename)
    stu_ws = stu_wb.active
    students = []
    for i in range(2, 500):
        name = stu_ws.cell(row=i, column=1).value
        idnum = stu_ws.cell(row=i, column=2).value
        if name == None:
            break
        students.append(Student(name, idnum))

    # Assign students to marked seats
    wb = load_workbook(seat_filename)
    idx = 0
    SEAT_RANGE = 200
    for ws in wb.worksheets:
        for i in range(1, SEAT_RANGE):
            for j in range(1, SEAT_RANGE):
                c = ws.cell(row=i, column=j)
                if c.value == 'X' or c.value == 'x':
                    if idx < len(students):
                        c.value = "{0}({1})".format(students[idx].name, students[idx].idnum)
                        idx += 1
                    else:
                        # clean X mark
                        c.value = None
    wb.save(output_filename)

if __name__ == '__main__':
    main()
